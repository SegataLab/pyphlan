#!/usr/bin/env python
#Author: Duy Tin Truong (duytin.truong@unitn.it)
#        at CIBIO, University of Trento, Italy


from openpyxl import Workbook
from openpyxl import load_workbook
import os

def table2xlsx(table, ofn, mode = 'write'):
	if mode == 'append' and os.path.exists(ofn):
		workbook = load_workbook(ofn)
		worksheet = workbook.create_sheet()
	else:
		workbook = Workbook()
		worksheet = workbook.active
	for r, row in enumerate(table): 
		for c, col in enumerate(row):
			worksheet.cell(row = r+1, column = c+1).value = col
	workbook.save(ofn)


